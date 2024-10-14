#!/usr/bin/env python
# -*- coding:UTF-8 -*-
#
# @AUTHOR: Rabbir
# @FILE: db.py
# @DATE: 2024/10/12
# @TIME: 16:50:01
#
# @DESCRIPTION: Excel 模块


import openpyxl


def get_file_path(file_name):
    """
    获取文件路径
    """
    try:
        with open("../{}".format(file_name), "r") as f:
            pass
        file_path = "../{}".format(file_name)
    except Exception as e:
        with open(file_name, "r") as f:
            pass
        file_path = "{}".format(file_name)
    print("Excel 文件路径为 {}\n".format(file_path) + "="*50)
    return file_path

def load_data(file_path):
    """
    读取 Excel 数据
    @return: {site_rows: [], spare_site_rows: []}
    """
    # 1. 加载 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    # 2. 读取网址表的数据
    ws = wb["网址"]
    site_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        # 如果第二列（自定义 ID）不为空
        if row[1] is not None:
            site_rows.append(row)
    # 3. 读取 备用链接地址（其他站点） 的数据
    ws = wb["备用链接地址（其他站点）"]
    spare_site_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        # 如果第一列（自定义 ID）不为空
        if row[0] is not None:
            spare_site_rows.append(row)
    # 4. 返回
    data = {
        "site_rows": site_rows,
        "spare_site_rows": spare_site_rows
    }
    print("Excel 中网址表总共有 {} 条数据（需要同步和不需要同步的都包含在内）".format(len(site_rows)))
    print("Excel 中备用链接地址（其他站点）表总共有 {} 条数据".format(len(spare_site_rows)))
    print("="*50)
    return data